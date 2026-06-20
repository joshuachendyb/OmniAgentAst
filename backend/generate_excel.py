import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

file_path = r"G:\OmniAgentAs-desk\backend\reports\2026年AI行业厂商对比分析.xlsx"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "AI厂商对比分析"

headers = ["企业名称", "所属领域", "核心产品/模型", "技术路线", "2025营收/估值(亿元)", "AI云市场份额", "核心优势", "主要劣势", "代表应用案例"]

header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
cell_align = Alignment(vertical="top", wrap_text=True)

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

data = [
    ["阿里巴巴", "综合AI", "通义千问(Qwen)大模型", "开源战略+B端深耕", "约9000(集团)", "约34%", "开源生态领先、代码生成能力强", "C端产品矩阵协同不足", "通义灵码、钉钉AI、通义万相"],
    ["字节跳动", "综合AI", "豆包、火山引擎、HiAgent 2.0", "应用反哺技术+C端驱动", "约1500(估值)", "约12%(AI云)", "产品化能力强、用户规模大", "底层算力依赖外部", "豆包AI助手、剪映AI、AI短剧"],
    ["腾讯", "综合AI", "混元大模型", "多模态+场景落地", "约6000(集团)", "约9%", "社交/游戏/广告场景丰富", "AI云份额下滑", "混元助手、腾讯会议AI、AI游戏"],
    ["华为", "综合AI", "盘古大模型、昇腾AI芯片", "软硬一体+全栈自研", "约6400(集团)", "约16%", "芯片自研、政企市场优势", "消费端生态较弱", "盘古农业/矿山模型、昇腾算力"],
    ["百度", "综合AI", "文心一言(ERNIE)", "全栈AI战略", "约1200(集团)", "约5%", "AI搜索+自动驾驶领先", "大模型生态相对封闭", "文心大模型、Apollo自动驾驶、AI搜索"],
    ["商汤科技", "计算机视觉/AI", "日日新(SenseNova)大模型", "AI基础设施+垂直应用", "约50(估值)", "约3%", "视觉技术领先、SenseCore平台", "盈利能力待验证", "AI医疗影像、智慧城市、AI创作"],
    ["科大讯飞", "语音AI/教育", "星火认知大模型", "语音技术+教育深耕", "约200", "约2%", "语音技术全球领先、教育壁垒", "通用大模型竞争激烈", "讯飞星火、AI学习机、智能录音笔"],
    ["MiniMax", "大模型/AI应用", "海螺AI、abab系列模型", "C端应用+大模型并行", "约100(估值)", "约1%", "产品迭代速度快", "算力资源有限", "海螺AI助手、AI角色扮演"],
    ["智谱AI", "大模型/AI", "GLM系列大模型", "学术背景+开源生态", "约80(估值)", "约1%", "清华背景、技术实力强", "商业化能力待提升", "智谱清言、CogVideo视频生成"],
    ["月之暗面(Kimi)", "大模型/AI", "Kimi大模型", "长上下文+AI搜索", "约150(估值)", "约1%", "长文本处理能力突出", "生态建设初期", "Kimi智能助手、AI搜索"],
    ["美团", "AI应用/本地生活", "AI调度系统", "场景驱动AI创新", "约1200(集团)", "不适用", "海量场景数据、实时调度能力", "非AI原生公司", "55毫秒路线规划、AI客服"],
    ["小米", "AI+硬件", "小爱同学、MIUI AI", "端侧AI+智能家居", "约3000(集团)", "不适用", "端侧部署能力、IoT生态", "云端算力相对薄弱", "小爱同学AI、AI拍照、智能汽车"],
    ["金山办公", "AI+办公", "WPS AI", "办公场景深耕", "约100", "不适用", "办公场景壁垒、用户量大", "大模型依赖外部", "WPS AI助手、AI文档生成"],
    ["旷视科技", "计算机视觉", "MegBrain、Face++", "AIoT+计算机视觉", "约30(估值)", "不适用", "视觉算法领先、物联网落地", "竞争加剧", "智慧物流、智慧零售、人脸支付"],
    ["云从科技", "人机协作AI", "从容大模型", "人机协同+行业解决方案", "约20(估值)", "不适用", "人机协同技术独特", "盈利压力大", "智慧金融、智慧治理、智慧商业"],
    ["商汤绝影", "自动驾驶AI", "SenseAuto", "车规级AI解决方案", "约50", "不适用", "车规级资质、全栈方案", "车企自研竞争", "智能座舱、自动驾驶辅助"],
    ["地平线", "AI芯片/自动驾驶", "征程系列芯片", "边缘AI芯片+自动驾驶", "约100(估值)", "不适用", "芯片自研、车企合作广泛", "生态建设周期长", "征程5/6芯片、智能驾驶方案"],
    ["寒武纪", "AI芯片", "思元系列AI芯片", "专用AI芯片设计", "约30(估值)", "不适用", "芯片性能领先、国产化替代", "生态建设不足", "思元590、边缘AI芯片"],
    ["阿里平头哥", "AI芯片", "含光系列AI芯片", "AI推理芯片自研", "不适用", "不适用", "芯片自研、阿里云协同", "对外销售有限", "含光800 AI推理芯片"],
    ["沐曦集成电路", "GPU芯片", "曦云GPU", "全栈GPU架构", "约20(估值)", "不适用", "团队技术背景强", "量产能力待验证", "训练/推理GPU芯片"],
    ["壁仞科技", "GPU芯片", "BR100系列GPU", "高性能GPU设计", "约150(估值)", "不适用", "芯片算力参数领先", "供应链制约", "BR100通用GPU芯片"],
    ["推想医疗", "AI+医疗", "AI医学影像诊断平台", "医学影像AI", "约5", "不适用", "FDA/NMPA认证", "市场规模有限", "肺结节CT辅助诊断"],
    ["同花顺", "AI+金融", "问财AI、iFinD", "金融AI应用", "约100", "不适用", "金融数据壁垒", "监管风险", "AI投顾、智能选股"],
    ["好未来", "AI+教育", "望道大模型、AI学习机", "教育AI应用", "约200", "不适用", "教育场景深耕", "政策环境影响", "AI自适应学习、智能批改"],
]

for row_idx, row_data in enumerate(data, 2):
    for col_idx, cell_value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
        cell.border = thin_border
        cell.alignment = cell_align

col_widths = [16, 18, 30, 28, 22, 18, 32, 32, 32]
for i, width in enumerate(col_widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

ws.auto_filter.ref = ws.dimensions
ws.freeze_panes = 'A2'

wb.save(file_path)
print(f"Excel文件已成功生成: {file_path}")
print(f"共写入{len(data)}条厂商数据")
