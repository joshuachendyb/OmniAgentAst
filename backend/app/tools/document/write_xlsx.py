# -*- coding: utf-8 -*-
# 编辑历史:
# 2026-06-22 - 小欧 - 创建文件，从document_tools.py拆分
# 2026-07-26 - 小欧 - summary加路径前空格
# 2026-07-31 - 小欧 - Bug⑳修复: data非数组或元素非dict时返回明确错误(原list[list]触发row.keys() AttributeError, 错误信息晦涩) | py_compile ✓
# 2026-08-07 - 小欧 - P04优化(北京老陈驱动 task001): 新增append_mode参数 — True=文件已存在时load_workbook末尾追加(表头一致性校验+按已有表头列序映射取值防串列), False=默认覆盖; else分支补mkdir防新建目录缺失崩溃; append分支error返回前补duration_ms计算防NameError | py_compile ✓
"""
D6: write_xlsx — 写入Excel文档

从document_tools.py拆分而来 — 小欧 2026-06-22
"""
# 【铁规1】helper/被调函数(以下划线_开头的函数)只返回raw dict，严禁调用build_success/build_error/build_warning和构建llm_data。
# build3+llm_data只能在tool的main函数(对外公开的函数)中包装。违反此规则的代码视为不合规。
# 【铁规2】工具返回原始data，禁止调用truncate_data_for_frontend。截断只能在前端yield层。
# 【铁规3】计时(duration_ms计算)只能在tool的主函数中，严禁在子函数/helper中计时。

import time as _time_mod
from pathlib import Path
from typing import Any, Dict, List, Optional  # 2026-07-31 小欧: 移除未使用 Union

from app.tools.tool_response import build_success, build_error
from app.tools.tool_fc_helper import _check_module
from app.tools.validate.file_type_checker import check_office_file
from app.tools.validate.file_safety_checker import check_content_safety
from app.tools.tool_constants import ERR_WRITE_XLSX  # 2026-07-31 小欧: 移除未使用 ERR_DOC_NO_OPENPYXL
from app.utils.json_utils import coerce_json
from app.tools.validate.file_path_checker import permission_error_hint, hint_for_write_error  # 2026-07-31 小欧: 移除未使用 logger
from app.utils.table_helper import get_table_header_style_config  # 2026-07-31 小欧: 移除未使用 calculate_column_widths


def _set_xlsx_table_style(ws):
    """设置Excel表格样式（表头背景色、数据单元格对齐和边框） — 小健 2026-06-24"""
    from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
    from openpyxl.utils import get_column_letter
    
    header_config = get_table_header_style_config()
    header_fill = PatternFill(
        start_color=header_config["bg_color"],
        end_color=header_config["bg_color"],
        fill_type="solid"
    )
    
    data_alignment = Alignment(horizontal="left", vertical="center")
    data_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col_idx in range(1, ws.max_column + 1):
        header_cell = ws.cell(row=1, column=col_idx)
        header_cell.fill = header_fill
        header_cell.font = Font(
            bold=header_config["bold"],
            color=header_config["text_color"]
        )
        
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = data_alignment
            cell.border = data_border


def _adjust_xlsx_column_width(ws):
    """调整Excel列宽自适应 — 小健 2026-06-24 — 小欧 2026-06-24 修复中文字符宽度"""
    from openpyxl.utils import get_column_letter
    
    def _display_width(s):
        """计算字符串显示宽度，中文字符占2宽度 — 小欧 2026-06-24"""
        width = 0
        for ch in str(s):
            if '\u4e00' <= ch <= '\u9fff' or '\u3000' <= ch <= '\u303f' or '\uff00' <= ch <= '\uffef':
                width += 2
            else:
                width += 1
        return width
    
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                max_len = max(max_len, _display_width(cell.value))
        ws.column_dimensions[col_letter].width = max(max_len + 2, 8)


def _build_write_xlsx_llm_data(
    exec_code: str, duration_ms: int,
    file_path: str = "", row_count: int = 0, detail: str = "",
    user_sheet_name: str = "", hint: str = "",
) -> Dict[str, Any]:
    """write_xlsx的llm_data构建函数 — 小欧 2026-06-22 — 小欧 2026-07-05 加hint参数"""
    _act_params = {"file_path": file_path}
    if user_sheet_name:
        _act_params["sheet_name"] = user_sheet_name
    if exec_code == "error":
        return {
            "summary": f"写入Excel {file_path}，失败: {detail}",
            "action": {"tool": "write_xlsx", "tool_zh": "写入Excel", "target": file_path, "params": _act_params},
            "status": {"exec_code": "error", "message": "写入Excel失败", "code": ERR_WRITE_XLSX, "detail": detail, "hint": hint if hint else "请检查路径和权限"},
            "duration_ms": duration_ms,
            "metrics": {},
        }
    return {
        "summary": f"写入Excel {file_path}，成功: {row_count}行",
        "action": {"tool": "write_xlsx", "tool_zh": "写入Excel", "target": file_path, "params": _act_params},
        "status": {"exec_code": "success", "message": "写入Excel成功", "code": "", "detail": "", "hint": ""},
        "duration_ms": duration_ms,
        "metrics": {
            "row_count": {"value": row_count, "text": f"{row_count}行"},
        },
    }


def write_xlsx(
    path: str,
    data: Optional[List[Dict[str, Any]]] = None,
    sheet_name: str = "Sheet1",
    append_mode: bool = False,
) -> Dict[str, Any]:
    """写入Excel文件 — 小沈 2026-06-16 — 小欧 2026-06-22 独立文件 — 小健 2026-06-24 参数简化 — 小欧 2026-06-24 增加文件类型前置检查 — 小欧 2026-08-07 新增append_mode追加模式"""
    t0 = _time_mod.perf_counter()

    # 文件类型前置检查（含路径检查+类型检查+模块安全检查）— 北京老陈 2026-07-09
    is_valid, error_detail, hint = check_office_file(path, allow_create=True)
    if not is_valid:
        duration_ms = int((_time_mod.perf_counter() - t0) * 1000)
        llm_data = _build_write_xlsx_llm_data("error", duration_ms, path, detail=error_detail, user_sheet_name=sheet_name, hint=hint)
        return build_error(data={}, llm_data=llm_data)

    data = coerce_json(data)
    # 2026-07-31 小欧: Bug⑳修复 — 明确data结构约束(list[dict]), 防list[list]触发row.keys() AttributeError(晦涩错误)
    if not isinstance(data, list):
        duration_ms = int((_time_mod.perf_counter() - t0) * 1000)
        llm_data = _build_write_xlsx_llm_data("error", duration_ms, path, detail=f"data参数必须是JSON数组, 收到类型: {type(data).__name__}", user_sheet_name=sheet_name, hint="data应为对象数组,如[{\"col1\":1,\"col2\":2}]")
        return build_error(data={}, llm_data=llm_data)
    for _i, _row in enumerate(data):
        if not isinstance(_row, dict):
            duration_ms = int((_time_mod.perf_counter() - t0) * 1000)
            llm_data = _build_write_xlsx_llm_data("error", duration_ms, path, detail=f"data的第{_i + 1}项必须是JSON对象, 收到类型: {type(_row).__name__}", user_sheet_name=sheet_name, hint="data的每个元素应为对象,如{\"col1\":1,\"col2\":2}")
            return build_error(data={}, llm_data=llm_data)
    cs_error, safe_data = check_content_safety(data, "xlsx", param_name="data")
    if cs_error:
        duration_ms = int((_time_mod.perf_counter() - t0) * 1000)
        llm_data = _build_write_xlsx_llm_data("error", duration_ms, path, detail=cs_error, user_sheet_name=sheet_name, hint=f"请检查data参数(当前类型: {type(data).__name__})")
        return build_error(data={}, llm_data=llm_data)
    data = safe_data

    if not _check_module("openpyxl"):
        duration_ms = int((_time_mod.perf_counter() - t0) * 1000)
        llm_data = _build_write_xlsx_llm_data("error", duration_ms, path, detail="openpyxl库未安装", user_sheet_name=sheet_name, hint="请安装openpyxl库")
        return build_error(data={}, llm_data=llm_data)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment

        headers = []
        rows = []
        if len(data) > 0:
            # KISS-DIRECT: 一行收集所有key，避免列不一致数据丢失 — 小健 2026-06-24
            headers = list(dict.fromkeys(k for row in data for k in row.keys()))
            # 按表头顺序填充，缺失填None
            rows = [[row.get(key) for key in headers] for row in data]

        wb = None
        ws = None
        _path = Path(path)
        if append_mode and _path.exists():
            from openpyxl import load_workbook
            wb = load_workbook(_path)
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
            # 表头一致性校验：已有表头缺失新列则拒绝（防静默错写）— 小欧 2026-08-07
            if headers and ws.max_row > 0:
                existing = [c.value for c in ws[1]]
                unknown = [h for h in headers if h not in existing]
                if unknown:
                    # error 返回前补 duration_ms 计算, 否则用未定义变量致 NameError — 小欧 2026-08-07
                    duration_ms = int((_time_mod.perf_counter() - t0) * 1000)
                    llm_data = _build_write_xlsx_llm_data("error", duration_ms, str(_path), detail=f"追加失败: 列 {unknown} 不在已有表头", user_sheet_name=sheet_name, hint="请保持 data 列与已有表头一致")
                    return build_error(data={}, llm_data=llm_data)
            # 按已有表头列序取值追加, 而非 rows(按 headers 顺序), 防止已有表头列序与 headers 不同导致串列 — 小欧 2026-08-07
            if rows:
                _cols = [c.value for c in ws[1]] if ws.max_row > 0 else headers
                for row_data in data:
                    ws.append([row_data.get(c) for c in _cols])
            # 追加后重新调整列宽适配新数据 — 小欧 2026-08-07
            _adjust_xlsx_column_width(ws)
            wb.save(_path)
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name

            if headers:
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center")

            if rows:
                for row_idx, row_data in enumerate(rows, 2):
                    for col_idx, cell_data in enumerate(row_data, 1):
                        ws.cell(row=row_idx, column=col_idx, value=cell_data)
            
            if headers or rows:
                _set_xlsx_table_style(ws)
                _adjust_xlsx_column_width(ws)

            # 新建文件前补 mkdir, 否则目标目录不存在时 wb.save 抛 FileNotFoundError — 小欧 2026-08-07
            _path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(_path)

        row_count = len(rows)
        duration_ms = int((_time_mod.perf_counter() - t0) * 1000)
        llm_data = _build_write_xlsx_llm_data("success", duration_ms, str(path), row_count, user_sheet_name=sheet_name)
        # =============================================================================
        # 数据设计：row_count/file_path 从 data 移除，通过 llm_data.metrics/summary
        # 传入 LLM observation。summary 已包含文件路径和行数：
        #   "写入Excel成功: /path.xlsx, 100行"
        # data 为空 dict 时 formatter 不追加详情，LLM 只看到 observation 行，
        # 避免 file_path 在 summary 和详情中重复造成冗余。
        # — 小欧 2026-07-06 18:46:13
        # =============================================================================
        return build_success(data={}, llm_data=llm_data)
    except PermissionError as e:
        duration_ms = int((_time_mod.perf_counter() - t0) * 1000)
        hint = permission_error_hint(path)
        # 小欧 2026-07-12: 此处path经Path()重赋值为WindowsPath,须str()化后传入builder,
        # 避免action.target持有Path对象触发观察格式化len()崩溃
        llm_data = _build_write_xlsx_llm_data("error", duration_ms, str(path), detail=str(e), user_sheet_name=sheet_name, hint=hint)
        return build_error(data={}, llm_data=llm_data)
    except Exception as e:
        duration_ms = int((_time_mod.perf_counter() - t0) * 1000)
        hint = hint_for_write_error(e, path)
        # 小欧 2026-07-12: 此处path经Path()重赋值为WindowsPath,须str()化后传入builder,
        # 避免action.target持有Path对象触发观察格式化len()崩溃
        llm_data = _build_write_xlsx_llm_data("error", duration_ms, str(path), detail=str(e), user_sheet_name=sheet_name, hint=hint)
        return build_error(data={}, llm_data=llm_data)
