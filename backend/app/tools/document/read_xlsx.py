# -*- coding: utf-8 -*-
# 编辑历史:
# 2026-07-15 - 小欧 - 常量归一化治理: 读取行数上限改引用 tool_constants.XLSX_MAX_ROWS(原硬编码10000), 功能零退化
# 2026-07-20 - 小欧 - 章15 门限治理:
#   1. XLSX_MAX_ROWS 依3.5改名
#     INER_READ_XLSX_MAX_ROWS(私有内部)
#   2. 3.4 硬安全防OOM: 触发置
#     data["truncated"]=True 附 reason
# 2026-07-20 - 小欧 - 修复:
#   1. sheet_name='None'字符串误判
#     为工作表名
#   2. 内函数WindowsPath泄漏行83/176 str()化
#   3. caller传str(path)合规标注行213/228
# 2026-07-21 - 小欧 - 字节安全治理: 移除行数限制(INER_READ_XLSX_MAX_ROWS), 
#    改为入口字节大小检查(READ_XLSX_INPUT_MAX_BYTES=20MB)防OOM; 
#    内部函数默认max_rows=1000000(安全兜底不再触发truncated)
# 2026-07-23 - 小欧 - 北京老陈驱动: 新增 Tool 层输出截断
#    rows 超 XLSX_OUTLIMIT_ROWS_MAX=1000 截断 + data.truncated=True
#    单格str超 XLSX_OUTLIMIT_CELL_CHARS=500 尾部截断
#    截断后缀包含原文长度: f"...(截断:原文N字符)"
#    row_count 同步更新, 使 llm_data.summary 与实际 data 一致
#     formatter #25 已读 truncated 字段, 自动显示 "⚠ 已截断"
# 2026-07-23 - 小欧 - 北京老陈驱动BugFix: 多 sheet 也 apply 行+格截断(bug4); 新增 truncated_reason 具体原因(bug5)
# 2026-07-24 - 小欧 - 修复: error summary嵌入full detail → 改用truncate_summary(detail)首行
# 2026-07-26 - 小欧 - OOD: 删 READ_XLSX_INPUT_MAX_BYTES 常量+入口检查, OOM自然抛出被except捕获(同dataanalysis模式)
# 2026-07-26 - 小欧 - 清理: 删logger死import(全文件无logger调用)
# 2026-07-26 - 小沈 - BugFix #2/#9: 更新stale docstring(删READ_XLSX_INPUT_MAX_BYTES引用); #3: path参数不覆盖
# 2026-08-13 - 小欧 - A5职责拆分: hint_* 错误提示函数/导入源改 app.tools.toolhelper.error_hints
"""
D4: read_xlsx — 读取Excel/CSV/XLS文档

从document_tools.py拆分而来 — 小欧 2026-06-22
内聚: _read_xlsx / _read_csv_stdlib 辅助函数 — 小欧 2026-06-24 移除._read_xls(不支持.xls)
"""
# 【铁规1】helper/被调函数(以下划线_开头的函数)只返回raw dict，严禁调用build_success/build_error/build_warning和构建llm_data。
# build3+llm_data只能在tool的main函数(对外公开的函数)中包装。违反此规则的代码视为不合规。
# 【铁规2】工具返回原始data，禁止调用truncate_data_for_frontend。截断只能在前端yield层。
# 【铁规3】计时(duration_ms计算)只能在tool的主函数中，严禁在子函数/helper中计时。
import csv
import time as _time_mod
from pathlib import Path
from typing import Any, Dict, Optional  # 2026-07-31 小欧: 移除未使用 List

from app.tools.tool_response import build_success, build_error
from app.tools.tool_fc_helper import _check_module
from app.tools.validate.file_type_checker import check_for_document_tool
from app.tools.toolhelper.error_hints import hint_for_read_error
from app.tools.tool_constants import (
    ERR_DOC_READ_XLSX,
    XLSX_OUTLIMIT_ROWS_MAX, XLSX_OUTLIMIT_CELL_CHARS,
)
from app.utils.text_utils import truncate_summary


def _build_read_xlsx_llm_data(
    exec_code: str, duration_ms: int,
    file_path: str = "", row_count: int = 0, sheet_count: int = 0, detail: str = "",
    user_sheet_name: str = "", hint: str = "",
) -> Dict[str, Any]:
    """read_xlsx的llm_data构建函数 — 小健 2026-06-21 — 小欧 2026-06-22 — 小欧 2026-07-05 加hint参数"""
    if exec_code == "error":
        _act_params = {"file_path": file_path}
        if user_sheet_name:
            _act_params["sheet_name"] = user_sheet_name
        _err_summary = truncate_summary(detail)
        return {
            "summary": f"读取Excel{file_path}，失败" + (f": {_err_summary}" if _err_summary else ""),
            "action": {"tool": "read_xlsx", "tool_zh": "读取Excel", "target": file_path, "params": _act_params},
            "status": {"exec_code": "error", "message": "读取Excel失败", "code": ERR_DOC_READ_XLSX, "detail": detail, "hint": hint if hint else "读取失败,详见错误明细"},
            "duration_ms": duration_ms,
            "metrics": {},
        }
    _act_params = {"file_path": file_path}
    if user_sheet_name:
        _act_params["sheet_name"] = user_sheet_name
    return {
        "summary": f"读取Excel{file_path}，成功: {row_count}行，{sheet_count}个工作表",
        "action": {"tool": "read_xlsx", "tool_zh": "读取Excel", "target": file_path, "params": _act_params},
        "status": {"exec_code": "success", "message": "读取Excel成功", "code": "", "detail": "", "hint": ""},
        "duration_ms": duration_ms,
        "metrics": {
            "row_count": {"value": row_count, "text": f"{row_count}行"},
            "sheet_count": {"value": sheet_count, "text": f"{sheet_count}个表"},
        },
    }


def _read_xlsx_inner(file_path: str, max_rows: int = 1000000, sheet_name: Optional[str] = None) -> Dict[str, Any]:
    """读取.xlsx文件(内部) — 小欧 2026-06-22
    OOM自然抛出被except捕获(同dataanalysis OOD模式) — 小沈 2026-07-26"""
    def _serialize_val(val):
        if val is None:
            return None
        if hasattr(val, 'isoformat'):
            return val.isoformat()
        return val

    from openpyxl import load_workbook

    path = Path(file_path)

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_names = wb.sheetnames
        if sheet_name and sheet_name not in sheet_names:
            return {"error_detail": f"工作表不存在: {sheet_name}", "hint": f"工作表 {sheet_name} 不存在,请确认工作表名称是否正确", "params": {"file_path": str(file_path), "sheet_name": sheet_name}}
        target_sheets = [sheet_name] if sheet_name else sheet_names

        all_sheets_data = []
        total_rows = 0
        for sheet in target_sheets:
            ws = wb[sheet]
            rows = []
            headers = []
            row_count = 0
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= max_rows:
                    break
                row_data = [_serialize_val(val) for val in row]
                if i == 0:
                    headers = [str(h) if h is not None else f"column_{j}" for j, h in enumerate(row_data)]
                else:
                    rows.append(row_data)
                    row_count += 1
            all_sheets_data.append({"sheet_name": sheet, "headers": headers, "rows": rows, "row_count": row_count})
            total_rows += row_count
    finally:
        wb.close()

    if len(all_sheets_data) == 1:
        result = all_sheets_data[0]
        result["sheet_names"] = sheet_names
        return result
    return {"sheets": all_sheets_data, "sheet_names": sheet_names, "row_count": total_rows}


def _read_csv_stdlib_inner(
    file_path: str,
    encoding: str = "utf-8",
    delimiter: str = ",",
    has_header: bool = True,
    max_rows: int = 1000000,
) -> Dict[str, Any]:
    """使用标准库csv读取CSV文件(内部) — 小欧 2026-06-22
    OOM自然抛出被except捕获(同dataanalysis OOD模式) — 小沈 2026-07-26"""
    path = Path(file_path)

    rows = []
    headers = []
    encodings_to_try = [encoding, "gbk", "gb2312", "latin-1"] if encoding == "utf-8" else [encoding, "utf-8", "latin-1"]
    read_ok = False
    for enc in encodings_to_try:
        try:
            with open(path, "r", encoding=enc, newline="") as f:
                reader = csv.reader(f, delimiter=delimiter)
                for i, row in enumerate(reader):
                    if i >= max_rows:
                        break
                    if i == 0:
                        if has_header:
                            headers = row
                        else:
                            headers = [f"col_{j}" for j in range(len(row))]
                            rows.append(row)
                    else:
                        rows.append(row)
            read_ok = True
            break
        except UnicodeDecodeError:
            continue
    if not read_ok:
        return {"error_detail": "编码不匹配", "hint": "无法以常见编码读取,请确认文件编码格式", "params": {"file_path": str(file_path), "encodings_tried": encodings_to_try}}

    return {"headers": headers, "rows": rows, "row_count": len(rows)}


def read_xlsx(path: str, sheet_name: Optional[str] = None) -> Dict[str, Any]:
    """读取Excel/CSV(.xlsx/.csv)文件 — 小沈 2026-06-19 — 小欧 2026-06-22 独立文件
    主函数: 负责build3+llm_data调用 — 小欧 2026-06-22
    参数: sheet_name - 指定工作表名（仅.xlsx），None则读取所有工作表 — 小健 2026-06-24
    小欧 2026-06-24 增加文件类型前置检查（.csv跳过检查） — 小欧 2026-06-24 移除.xls死代码
    2026-07-20 小欧: 兼容LLM传入sheet_name='None'等字符串"""
    if isinstance(sheet_name, str) and sheet_name.strip().lower() in ("none", "null", ""):
        sheet_name = None
    _p = Path(path)
    suffix = _p.suffix.lower()
    t0 = _time_mod.perf_counter()

    # 文件类型前置检查（.csv由本工具处理，跳过检查） — 小欧 2026-06-24
    if suffix != ".csv":
        is_valid, error_detail, suggested_tool = check_for_document_tool(str(path))
        if not is_valid:
            duration_ms = int((_time_mod.perf_counter() - t0) * 1000)
            _sn = sheet_name or ""
            if suggested_tool:
                _hint = f"建议使用{suggested_tool}工具"
            elif suggested_tool == "":
                _hint = "请检查文件路径和文件名是否正确"
            else:
                _hint = "文件类型不匹配,请使用.xlsx或.csv格式"
            llm_data = _build_read_xlsx_llm_data("error", duration_ms, path, detail=error_detail, user_sheet_name=_sn, hint=_hint)
            return build_error(data={}, llm_data=llm_data)

    if suffix == ".csv":
        try:
            result = _read_csv_stdlib_inner(str(path), encoding="utf-8", delimiter=",", has_header=True)
        except Exception as e:
            duration_ms = int((_time_mod.perf_counter() - t0) * 1000)
            llm_data = _build_read_xlsx_llm_data("error", duration_ms, path, detail=str(e), user_sheet_name=sheet_name or "", hint=hint_for_read_error(e, path))
            return build_error(data={}, llm_data=llm_data)
    else:
        if not _check_module("openpyxl"):
            duration_ms = int((_time_mod.perf_counter() - t0) * 1000)
            llm_data = _build_read_xlsx_llm_data("error", duration_ms, path, detail="openpyxl库未安装", user_sheet_name=sheet_name or "", hint="请安装openpyxl库")
            return build_error(data={}, llm_data=llm_data)
        try:
            result = _read_xlsx_inner(str(path), sheet_name=sheet_name)
        except Exception as e:
            duration_ms = int((_time_mod.perf_counter() - t0) * 1000)
            llm_data = _build_read_xlsx_llm_data("error", duration_ms, path, detail=str(e), user_sheet_name=sheet_name or "", hint=hint_for_read_error(e, path))
            return build_error(data={}, llm_data=llm_data)

    duration_ms = int((_time_mod.perf_counter() - t0) * 1000)
    if "error_detail" in result:
        detail = result["error_detail"]
        llm_data = _build_read_xlsx_llm_data("error", duration_ms, path, detail=detail, user_sheet_name=sheet_name or "", hint=result.get("hint", ""))
        return build_error(data=result, llm_data=llm_data)
    else:
        row_count = result.get("row_count", 0)
        sheet_count = len(result.get("sheet_names", []))
        result.pop("row_count", None)
        # ── Tool 层输出截断 — 小欧 2026-07-23 — 小欧 2026-07-23 BugFix: 多 sheet 截断+truncated_reason
        # 行数截断: 超 XLSX_OUTLIMIT_ROWS_MAX 截断 + data.truncated=True
        # 单格字符串截断: 超 XLSX_OUTLIMIT_CELL_CHARS 尾部加 "...(截断:原文N字符)"
        # formatter #25 读 data.truncated 自动显示 "⚠ 已截断"
        # 注意: row_count 同步更新, 使 llm_data.summary 与实际 data 一致
        _tool_trunc = False
        _trunc_rows = False
        _trunc_cells = False

        def _truncate_rows_and_cells(rows_list: list) -> list:
            """辅助函数: 行截断(超XLSX_OUTLIMIT_ROWS_MAX) + 格截断(超XLSX_OUTLIMIT_CELL_CHARS) — 小欧 2026-07-23"""
            nonlocal _tool_trunc, _trunc_rows, _trunc_cells
            if len(rows_list) > XLSX_OUTLIMIT_ROWS_MAX:
                rows_list = rows_list[:XLSX_OUTLIMIT_ROWS_MAX]
                _tool_trunc = _trunc_rows = True
            if rows_list:
                _new = []
                for row in rows_list:
                    _nr = []
                    for v in row:
                        if isinstance(v, str) and len(v) > XLSX_OUTLIMIT_CELL_CHARS:
                            _nr.append(v[:XLSX_OUTLIMIT_CELL_CHARS] + f"...(截断:原文{len(v)}字符)")
                            _tool_trunc = _trunc_cells = True
                        else:
                            _nr.append(v)
                    _new.append(_nr)
                rows_list = _new
            return rows_list

        if "sheets" in result:
            # 多 sheet: 每张表独立截断
            for s in result["sheets"]:
                s["rows"] = _truncate_rows_and_cells(s.get("rows", []))
                s["row_count"] = len(s["rows"])
            row_count = sum(s.get("row_count", 0) for s in result["sheets"])
        else:
            # 单 sheet / CSV
            result["rows"] = _truncate_rows_and_cells(result.get("rows", []))
            if _tool_trunc:
                row_count = min(row_count, len(result["rows"]))

        if _tool_trunc:
            result["truncated"] = True
            _reason = []
            if _trunc_rows: _reason.append(f"行数超{XLSX_OUTLIMIT_ROWS_MAX}")
            if _trunc_cells: _reason.append(f"单格字符超{XLSX_OUTLIMIT_CELL_CHARS}")
            result["truncated_reason"] = "、".join(_reason)
        # ── ──
        llm_data = _build_read_xlsx_llm_data("success", duration_ms, path, row_count, sheet_count, user_sheet_name=sheet_name or "")
        # =============================================================================
        # 数据设计：row_count/sheet_count 从 data 移除，通过 llm_data.metrics 传入 summary
        # summary 示例: "读取Excel成功: 100行, 3个工作表"
        # — 小欧 2026-07-06 18:46:13
        # =============================================================================
        # ---- observation_formatter route -------------------------------------------
        # branch: #25 read_xlsx 专属(单sheet/CSV headers+rows) / #21 scalar fallback(多sheet)
        # trigger: action.tool=="read_xlsx" 且 "headers" in data and "rows" in data
        # handler: _format_xlsx_result(data, llm_data) — 专属 #25 全量展示(无显示域行/列截断, read_xlsx 无offset分页); 两态说明仅反映 Tool 层 truncated
        # note:    多sheet返回 {"sheets": [...], "sheet_names": [...]}, 无headers/rows, 走 #21 fallback
        # file:    observation_formatter.py
        # ------------------------------------------------------------------------------
        return build_success(data=result, llm_data=llm_data)
