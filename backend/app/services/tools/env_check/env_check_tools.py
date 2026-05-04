# -*- coding: utf-8 -*-
"""
环境检查工具函数模块

【创建时间】2026-05-02 小沈
【设计依据】按文档第8.4节 Tool 83-91 定义

【重要】新函数增加规范 - 小沈 2026-05-04
新增函数时必须同步修改以下3个文件：
1. *_tools.py: 函数实现（必须有详细注释）
2. *_schema.py: Pydantic 模型（输入参数定义）
3. *_register.py: 显式注册（description + examples + input_model）

包含9个工具：
- check_python_available: 检查Python环境
- validate_code_safety: 验证代码安全性
- check_node_available: 检查Node.js环境
- check_module_available: 检查Python模块
- validate_csv_format: 验证CSV文件格式
- validate_chart_data: 验证图表数据格式
- check_pdf_readable: 检查PDF可读性
- check_docx_readable: 检查Word可读性
- check_xlsx_readable: 检查Excel可读性

【2026-05-02 小沈重构】
- 移除所有 @register_tool 装饰器，改由 env_check_register.py 显式注册

Author: 小沈 - 2026-05-02
"""

import os
import re
import sys
import csv
import importlib
import subprocess
from typing import Dict, Any
from pathlib import Path


DANGEROUS_PATTERNS = [
    (r"os\.system\s*\(", "系统调用(os.system)"),
    (r"subprocess\.(call|run|Popen|check_output)\s*\(", "子进程调用(subprocess)"),
    (r"shutil\.rmtree\s*\(", "递归删除目录(shutil.rmtree)"),
    (r"os\.remove\s*\(", "删除文件(os.remove)"),
    (r"os\.unlink\s*\(", "删除文件(os.unlink)"),
    (r"__import__\s*\(", "动态导入(__import__)"),
    (r"eval\s*\(", "动态执行(eval)"),
    (r"exec\s*\(", "动态执行(exec)"),
    (r"compile\s*\(", "动态编译(compile)"),
    (r"open\s*\(.*[\'\"]w[\'\"]", "写入文件操作"),
    (r"socket\s*\.", "网络Socket操作"),
    (r"requests\.(get|post|put|delete|patch)\s*\(", "HTTP请求(requests)"),
    (r"urllib\.request", "URL请求(urllib)"),
]


def check_python_available() -> Dict[str, Any]:
    """检查Python环境是否可用 - 小沈 2026-05-02"""
    try:
        version = sys.version
        executable = sys.executable
        return {
            "code": "SUCCESS",
            "data": {
                "available": True,
                "version": version,
                "executable": executable,
            },
            "message": f"Python环境可用，版本: {version}"
        }
    except Exception as e:
        return {
            "code": "SUCCESS",
            "data": {"available": False, "version": None, "executable": None},
            "message": f"Python环境检查失败: {str(e)}"
        }


def validate_code_safety(code: str) -> Dict[str, Any]:
    """验证代码安全性 - 小沈 2026-05-02"""
    warnings = []
    for pattern, desc in DANGEROUS_PATTERNS:
        if re.search(pattern, code):
            warnings.append(desc)

    is_safe = len(warnings) == 0

    return {
        "code": "SUCCESS",
        "data": {
            "safe": is_safe,
            "warnings": warnings,
            "warning_count": len(warnings),
        },
        "message": "代码安全" if is_safe else f"代码存在 {len(warnings)} 个安全风险: {', '.join(warnings)}"
    }


def check_node_available() -> Dict[str, Any]:
    """检查Node.js环境是否可用 - 小沈 2026-05-02"""
    try:
        result = subprocess.run(
            ["node", "--version"],
            capture_output=True,
            text=True,
            timeout=5
        )
        if result.returncode == 0:
            version = result.stdout.strip()
            return {
                "code": "SUCCESS",
                "data": {"available": True, "version": version},
                "message": f"Node.js环境可用，版本: {version}"
            }
        else:
            return {
                "code": "SUCCESS",
                "data": {"available": False, "version": None},
                "message": "Node.js环境不可用"
            }
    except FileNotFoundError:
        return {
            "code": "SUCCESS",
            "data": {"available": False, "version": None},
            "message": "Node.js未安装"
        }
    except Exception as e:
        return {
            "code": "SUCCESS",
            "data": {"available": False, "version": None},
            "message": f"Node.js环境检查失败: {str(e)}"
        }


def check_module_available(module_name: str) -> Dict[str, Any]:
    """检查Python模块是否已安装 - 小沈 2026-05-02"""
    try:
        mod = importlib.import_module(module_name)
        version = getattr(mod, "__version__", "unknown")
        return {
            "code": "SUCCESS",
            "data": {"available": True, "version": version},
            "message": f"模块 {module_name} 已安装，版本: {version}"
        }
    except ImportError:
        return {
            "code": "SUCCESS",
            "data": {"available": False, "version": None},
            "message": f"模块 {module_name} 未安装"
        }
    except Exception as e:
        return {
            "code": "SUCCESS",
            "data": {"available": False, "version": None},
            "message": f"模块检查失败: {str(e)}"
        }


def validate_csv_format(file_path: str) -> Dict[str, Any]:
    """验证CSV文件格式 - 小沈 2026-05-02"""
    path = Path(file_path)
    if not path.exists():
        return {
            "code": "SUCCESS",
            "data": {"valid": False, "errors": ["文件不存在"]},
            "message": f"文件不存在: {file_path}"
        }

    if not path.suffix.lower() in (".csv", ".tsv", ".txt"):
        return {
            "code": "SUCCESS",
            "data": {"valid": False, "errors": ["文件扩展名不是CSV格式"]},
            "message": "文件扩展名不是CSV格式"
        }

    errors = []
    try:
        with open(path, "r", encoding="utf-8", newline="") as f:
            reader = csv.reader(f)
            row_count = 0
            col_count = None
            for i, row in enumerate(reader):
                row_count += 1
                if col_count is None:
                    col_count = len(row)
                elif len(row) != col_count:
                    if len(row) != 0:
                        errors.append(f"第{i+1}行列数不一致: 期望{col_count}列，实际{len(row)}列")
                if row_count > 1000:
                    break
    except UnicodeDecodeError:
        try:
            with open(path, "r", encoding="gbk", newline="") as f:
                reader = csv.reader(f)
                row_count = 0
                for _ in reader:
                    row_count += 1
                    if row_count > 1000:
                        break
        except Exception as e:
            errors.append(f"文件编码错误: {str(e)}")
    except Exception as e:
        errors.append(f"文件读取错误: {str(e)}")

    is_valid = len(errors) == 0
    return {
        "code": "SUCCESS",
        "data": {"valid": is_valid, "errors": errors},
        "message": "CSV格式正确" if is_valid else f"CSV格式有 {len(errors)} 个问题"
    }


def validate_chart_data(data: Dict[str, Any]) -> Dict[str, Any]:
    """验证图表数据格式 - 小沈 2026-05-02"""
    errors = []

    if not isinstance(data, dict):
        errors.append("data必须是字典类型")
    else:
        if "labels" not in data:
            errors.append("缺少labels字段")
        elif not isinstance(data["labels"], list):
            errors.append("labels必须是数组类型")

        if "values" not in data:
            errors.append("缺少values字段")
        elif not isinstance(data["values"], list):
            errors.append("values必须是数组类型")

        if "labels" in data and "values" in data:
            if isinstance(data["labels"], list) and isinstance(data["values"], list):
                if len(data["labels"]) != len(data["values"]):
                    errors.append(f"labels和values长度不一致: labels={len(data['labels'])}, values={len(data['values'])}")

    is_valid = len(errors) == 0
    return {
        "code": "SUCCESS",
        "data": {"valid": is_valid, "errors": errors},
        "message": "图表数据格式正确" if is_valid else f"图表数据格式有 {len(errors)} 个问题"
    }


def check_pdf_readable(file_path: str) -> Dict[str, Any]:
    """检查PDF文件是否可读 - 小沈 2026-05-02"""
    path = Path(file_path)
    if not path.exists():
        return {
            "code": "SUCCESS",
            "data": {"readable": False, "error": "文件不存在"},
            "message": f"文件不存在: {file_path}"
        }

    try:
        import pdfplumber
        with pdfplumber.open(path) as pdf:
            page_count = len(pdf.pages)
        return {
            "code": "SUCCESS",
            "data": {"readable": True, "page_count": page_count},
            "message": f"PDF文件可读，共 {page_count} 页"
        }
    except ImportError:
        return {
            "code": "SUCCESS",
            "data": {"readable": False, "error": "pdfplumber库未安装"},
            "message": "pdfplumber库未安装，无法检查PDF文件"
        }
    except Exception as e:
        return {
            "code": "SUCCESS",
            "data": {"readable": False, "error": str(e)},
            "message": f"PDF文件不可读: {str(e)}"
        }


def check_docx_readable(file_path: str) -> Dict[str, Any]:
    """检查Word文件是否可读 - 小沈 2026-05-02"""
    path = Path(file_path)
    if not path.exists():
        return {
            "code": "SUCCESS",
            "data": {"readable": False, "error": "文件不存在"},
            "message": f"文件不存在: {file_path}"
        }

    try:
        import docx
        doc = docx.Document(path)
        para_count = len(doc.paragraphs)
        return {
            "code": "SUCCESS",
            "data": {"readable": True, "paragraph_count": para_count},
            "message": f"Word文件可读，共 {para_count} 段"
        }
    except ImportError:
        return {
            "code": "SUCCESS",
            "data": {"readable": False, "error": "python-docx库未安装"},
            "message": "python-docx库未安装，无法检查Word文件"
        }
    except Exception as e:
        return {
            "code": "SUCCESS",
            "data": {"readable": False, "error": str(e)},
            "message": f"Word文件不可读: {str(e)}"
        }


def check_xlsx_readable(file_path: str) -> Dict[str, Any]:
    """检查Excel文件是否可读 - 小沈 2026-05-02"""
    path = Path(file_path)
    if not path.exists():
        return {
            "code": "SUCCESS",
            "data": {"readable": False, "error": "文件不存在"},
            "message": f"文件不存在: {file_path}"
        }

    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True)
        sheet_names = wb.sheetnames
        wb.close()
        return {
            "code": "SUCCESS",
            "data": {"readable": True, "sheet_names": sheet_names},
            "message": f"Excel文件可读，共 {len(sheet_names)} 个工作表"
        }
    except ImportError:
        return {
            "code": "SUCCESS",
            "data": {"readable": False, "error": "openpyxl库未安装"},
            "message": "openpyxl库未安装，无法检查Excel文件"
        }
    except Exception as e:
        return {
            "code": "SUCCESS",
            "data": {"readable": False, "error": str(e)},
            "message": f"Excel文件不可读: {str(e)}"
        }
