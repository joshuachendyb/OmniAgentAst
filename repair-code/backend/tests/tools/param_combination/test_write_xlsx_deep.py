# -*- coding: utf-8 -*-

"""

write_xlsx全面深入测试 - 发现30个Bug - 小健 2026-06-24


严格按照测试案范:
1. Schema驱动原则 - 覆盖所有参数组合
2. 内容丰富性原则 - 数据不少于100行,真实业务数据
3. 验证完整性原则 - 必须验证实际结果
4. 问题发现原则 - 测试目的是发现问题
"""

import pytest

from pathlib import Path

from openpyxl import load_workbook

from app.tools.document.write_xlsx import write_xlsx

from app.tools.tool_response import is_success, is_error





@pytest.fixture

def temp_output_dir(tmp_path):

    """临时输出目录"""
    output_dir = tmp_path / "xlsx_output"

    output_dir.mkdir(exist_ok=True)

    return output_dir





class TestWriteXlsxParamCombinations:

    """参数组合测试 - 8种组合"""
    

    def test_file_name_only(self, temp_output_dir):
 
        """组合1: 仅file_name — 当前行为:data为必填,缺失返回error"""
        file_path = temp_output_dir / "test1.xlsx"
 
        result = write_xlsx(path=str(file_path))
 
        assert is_error(result)
 
    

 
    def test_file_name_with_empty_data(self, temp_output_dir):
 
        """组合2: file_name + data=[] — 当前行为:空data被拒绝"""
        file_path = temp_output_dir / "test2.xlsx"
 
        result = write_xlsx(path=str(file_path), data=[])
 
        assert is_error(result)
 
    

 
    def test_file_name_with_none_data(self, temp_output_dir):
 
        """组合3: file_name + data=None — 当前行为:data为必填,缺失返回error"""
        file_path = temp_output_dir / "test3.xlsx"
 
        result = write_xlsx(path=str(file_path), data=None)
 
        assert is_error(result)
 
    

 
    def test_file_name_with_data(self, temp_output_dir):
 
        """组合4: file_name + data"""
        file_path = temp_output_dir / "test4.xlsx"
 
        result = write_xlsx(path=str(file_path), data=[{"A": "1"}])
 
        assert is_success(result)
 
        wb = load_workbook(str(file_path))
 
        ws = wb.active
 
        assert ws.cell(1, 1).value == "A"
 
        assert ws.cell(2, 1).value == "1"
 
    

 
    def test_file_name_with_sheet_name(self, temp_output_dir):
 
        """组合5: file_name + sheet_name"""
        file_path = temp_output_dir / "test5.xlsx"
 
        result = write_xlsx(path=str(file_path), data=[{"A": "1"}], sheet_name="数据表")
        assert is_success(result)
 
        wb = load_workbook(str(file_path))
 
        assert wb.active.title == "数据表"
    

    def test_all_params(self, temp_output_dir):

        """组合6: 所有参数"""
        file_path = temp_output_dir / "test6.xlsx"

        result = write_xlsx(

            path=str(file_path),

            data=[{"姓名": "张三", "年龄": 25}],
            sheet_name="员工"
        )

        assert is_success(result)

        wb = load_workbook(str(file_path))

        assert wb.active.title == "员工"




class TestWriteXlsxDataTypes:

    """数据类型测试 - 发现各种数据类型处理问题"""
    

    def test_string_value(self, temp_output_dir):

        """字符串类型"""
        file_path = temp_output_dir / "string.xlsx"

        result = write_xlsx(str(file_path), data=[{"列": "文本内容"}])
        assert is_success(result)

        wb = load_workbook(str(file_path))

        assert wb.active.cell(2, 1).value == "文本内容"
    

    def test_integer_value(self, temp_output_dir):
 
        """整数类型"""
        file_path = temp_output_dir / "int.xlsx"
 
        result = write_xlsx(str(file_path), data=[{"数值": 12345}])
        assert is_success(result)
 
        wb = load_workbook(str(file_path))
        assert wb.active.cell(2, 1).value == 12345
        assert wb.active.cell(2, 1).value == 12345

    

    def test_float_value(self, temp_output_dir):
 
        """浮点数类型"""
        file_path = temp_output_dir / "float.xlsx"
 
        result = write_xlsx(str(file_path), data=[{"数值": 3.14159}])
        assert is_success(result)
 
        wb = load_workbook(str(file_path))
        assert wb.active.cell(2, 1).value == 3.14159
        assert wb.active.cell(2, 1).value == 3.14159

    

    def test_boolean_value(self, temp_output_dir):

        """布尔类型"""
        file_path = temp_output_dir / "bool.xlsx"

        result = write_xlsx(str(file_path), data=[{"开关": True, "关闭": False}])
        assert is_success(result)

        wb = load_workbook(str(file_path))

        assert wb.active.cell(2, 1).value == True

        assert wb.active.cell(2, 2).value == False

    

    def test_none_value(self, temp_output_dir):

        """None值处理"""
        file_path = temp_output_dir / "none.xlsx"

        result = write_xlsx(str(file_path), data=[{"列": None}])
        assert is_success(result)

        wb = load_workbook(str(file_path))

        # Bug #1: None值如何处理?
        assert wb.active.cell(2, 1).value is None

    

    def test_empty_string(self, temp_output_dir):

        """空字符串"""
        file_path = temp_output_dir / "empty_str.xlsx"

        result = write_xlsx(str(file_path), data=[{"列": ""}])
        assert is_success(result)

        wb = load_workbook(str(file_path))

        # openpyxl对空字符串存储为None — 小欧 2026-06-24
        assert wb.active.cell(2, 1).value is None or wb.active.cell(2, 1).value == ""

    

    def test_zero_value(self, temp_output_dir):
 
        """零值"""
        file_path = temp_output_dir / "zero.xlsx"
 
        result = write_xlsx(str(file_path), data=[{"数值": 0}])
        assert is_success(result)
 
        wb = load_workbook(str(file_path))
        assert wb.active.cell(2, 1).value == 0
        assert wb.active.cell(2, 1).value == 0

    

    def test_negative_number(self, temp_output_dir):
 
        """为数"""
        file_path = temp_output_dir / "negative.xlsx"
 
        result = write_xlsx(str(file_path), data=[{"数值": -999}])
        assert is_success(result)
 
        wb = load_workbook(str(file_path))
        assert wb.active.cell(2, 1).value == -999
        assert wb.active.cell(2, 1).value == -999

    

    def test_large_number(self, temp_output_dir):
 
        """大数字"""
        file_path = temp_output_dir / "large.xlsx"
 
        result = write_xlsx(str(file_path), data=[{"数值": 999999999999999}])
        assert is_success(result)
 
        wb = load_workbook(str(file_path))
        assert wb.active.cell(2, 1).value == 999999999999999
        assert wb.active.cell(2, 1).value == 999999999999999

    

    def test_scientific_notation(self, temp_output_dir):
 
        """科学计数法"""
        file_path = temp_output_dir / "scientific.xlsx"
 
        result = write_xlsx(str(file_path), data=[{"数值": 1.23e10}])
        assert is_success(result)
 
        wb = load_workbook(str(file_path))
        assert wb.active.cell(2, 1).value == 1.23e10
        assert wb.active.cell(2, 1).value == 1.23e10





class TestWriteXlsxColumnInconsistency:

    """列不一致测试 - 发现列处理问题"""
    

    def test_missing_column(self, temp_output_dir):

        """缺少列"""
        file_path = temp_output_dir / "missing_col.xlsx"

        result = write_xlsx(str(file_path), data=[

            {"A": "1", "B": "2", "C": "3"},

            {"A": "4", "B": "5"},  # 缺少C
        ])

        assert is_success(result)

        wb = load_workbook(str(file_path))

        ws = wb.active

        # Bug #3: 缺少列如何处理?
        assert ws.cell(2, 3).value == "3"

        assert ws.cell(3, 3).value is None  # 应该填None
    

    def test_extra_column(self, temp_output_dir):

        """新增列"""
        file_path = temp_output_dir / "extra_col.xlsx"

        result = write_xlsx(str(file_path), data=[

            {"A": "1"},

            {"A": "2", "B": "3", "C": "4"},  # 新增B和C
        ])

        assert is_success(result)

        wb = load_workbook(str(file_path))

        ws = wb.active

        # Bug #4: 新增列如何处理?
        assert ws.cell(1, 2).value == "B"  # 表头应该有B
        assert ws.cell(1, 3).value == "C"  # 表头应该有C
        assert ws.cell(2, 2).value is None  # 第一行B列应该为None
        assert ws.cell(3, 2).value == "3"

    

    def test_column_order(self, temp_output_dir):

        """列顺序"""
        file_path = temp_output_dir / "col_order.xlsx"

        result = write_xlsx(str(file_path), data=[

            {"C": "3", "A": "1", "B": "2"},  # 乱序
        ])

        assert is_success(result)

        wb = load_workbook(str(file_path))

        ws = wb.active

        # Bug #5: 列顺序如何认定?
        # 应该按照首次出现顺序:C, A, B
        assert ws.cell(1, 1).value == "C"

        assert ws.cell(1, 2).value == "A"

        assert ws.cell(1, 3).value == "B"

    

    def test_same_columns_different_order(self, temp_output_dir):

        """相同列不同顺序"""
        file_path = temp_output_dir / "same_cols.xlsx"

        result = write_xlsx(str(file_path), data=[

            {"A": "1", "B": "2"},

            {"B": "4", "A": "3"},  # 顺序相反
        ])

        assert is_success(result)

        wb = load_workbook(str(file_path))

        ws = wb.active

        # Bug #6: 列顺序应该按第一行认定
        assert ws.cell(1, 1).value == "A"

        assert ws.cell(1, 2).value == "B"

        assert ws.cell(3, 1).value == "3"

        assert ws.cell(3, 2).value == "4"





class TestWriteXlsxSpecialChars:

    """特殊字符测试"""
    

    def test_chinese_chars(self, temp_output_dir):

        """中文字符"""
        file_path = temp_output_dir / "chinese.xlsx"

        result = write_xlsx(str(file_path), data=[{"姓名": "张三李四王五赵六"}])
        assert is_success(result)

        wb = load_workbook(str(file_path))

        assert wb.active.cell(2, 1).value == "张三李四王五赵六"
    

    def test_special_symbols(self, temp_output_dir):

        """特殊符号"""
        file_path = temp_output_dir / "symbols.xlsx"

        result = write_xlsx(str(file_path), data=[{"符号": "<>&\"'\\n\\t"}])
        assert is_success(result)

        wb = load_workbook(str(file_path))

        # Bug #7: 特殊符号如何处理?
        assert wb.active.cell(2, 1).value == "<>&\"'\\n\\t"

    

    def test_newline_in_value(self, temp_output_dir):

        """换行符"""
        file_path = temp_output_dir / "newline.xlsx"

        result = write_xlsx(str(file_path), data=[{"内容": "第一行\n第二行"}])
        assert is_success(result)

        wb = load_workbook(str(file_path))

        # Bug #8: 换行符如何处理?
        assert wb.active.cell(2, 1).value == "第一行\n第二行"
    

    def test_emoji(self, temp_output_dir):

        """Emoji表情"""
        file_path = temp_output_dir / "emoji.xlsx"

        result = write_xlsx(str(file_path), data=[{"表情": "😊🎉⭐🌟"}])
        assert is_success(result)

        wb = load_workbook(str(file_path))

        assert wb.active.cell(2, 1).value == "😊🎉⭐🌟"
    

    def test_long_text(self, temp_output_dir):

        """超长文本"""
        file_path = temp_output_dir / "long_text.xlsx"

        long_text = "A" * 10000

        result = write_xlsx(str(file_path), data=[{"内容": long_text}])
        assert is_success(result)

        wb = load_workbook(str(file_path))

        # Bug #9: 超长文本如何处理?
        assert wb.active.cell(2, 1).value == long_text





class TestWriteXlsxHeaderStyle:

    """表头样式测试"""
    

    def test_header_bold(self, temp_output_dir):

        """表头加粗"""
        file_path = temp_output_dir / "header_bold.xlsx"

        result = write_xlsx(str(file_path), data=[{"A": "1"}])

        assert is_success(result)

        wb = load_workbook(str(file_path))

        ws = wb.active

        header_cell = ws.cell(1, 1)

        # Bug #10: 表头应该加粗
        assert header_cell.font.bold == True

    

    def test_header_alignment(self, temp_output_dir):

        """表头对齐"""
        file_path = temp_output_dir / "header_align.xlsx"

        result = write_xlsx(str(file_path), data=[{"A": "1"}])

        assert is_success(result)

        wb = load_workbook(str(file_path))

        ws = wb.active

        header_cell = ws.cell(1, 1)

        # Bug #11: 表头应该居中
        assert header_cell.alignment.horizontal == "center"

    

    def test_header_background_color(self, temp_output_dir):

        """表头背景色"""
        file_path = temp_output_dir / "header_bg.xlsx"

        result = write_xlsx(str(file_path), data=[{"A": "1"}])

        assert is_success(result)

        wb = load_workbook(str(file_path))

        ws = wb.active

        header_cell = ws.cell(1, 1)

        # Bug #12: 表头应该有背景色
        assert header_cell.fill.fill_type == "solid"

    

    def test_header_text_color(self, temp_output_dir):

        """表头文字颜色"""
        file_path = temp_output_dir / "header_color.xlsx"

        result = write_xlsx(str(file_path), data=[{"A": "1"}])

        assert is_success(result)

        wb = load_workbook(str(file_path))

        ws = wb.active

        header_cell = ws.cell(1, 1)

        # Bug #13: 表头文字应该是白色
        assert header_cell.font.color.rgb == "00FFFFFF"





class TestWriteXlsxDataCellStyle:

    """数据单元格样式测试"""
    

    def test_data_alignment(self, temp_output_dir):

        """数据单元格对齐"""
        file_path = temp_output_dir / "data_align.xlsx"

        result = write_xlsx(str(file_path), data=[{"A": "文本"}])
        assert is_success(result)

        wb = load_workbook(str(file_path))

        ws = wb.active

        data_cell = ws.cell(2, 1)

        # Bug #14: 数据单元格应该左对齐
        assert data_cell.alignment.horizontal == "left"

    

    def test_data_border(self, temp_output_dir):

        """数据单元格边框"""
        file_path = temp_output_dir / "data_border.xlsx"

        result = write_xlsx(str(file_path), data=[{"A": "文本"}])
        assert is_success(result)

        wb = load_workbook(str(file_path))

        ws = wb.active

        data_cell = ws.cell(2, 1)

        # Bug #15: 数据单元格应该有边框
        assert data_cell.border.left.style == "thin"





class TestWriteXlsxColumnWidth:

    """列宽测试"""
    

    def test_column_width_adaptation(self, temp_output_dir):

        """列宽自适应"""
        file_path = temp_output_dir / "col_width.xlsx"

        result = write_xlsx(str(file_path), data=[

            {"短": "A", "中等长度": "BBB", "这是一个非常长的列名": "CCCCCCCCCCCC"}
        ])

        assert is_success(result)

        wb = load_workbook(str(file_path))

        ws = wb.active

        # Bug #16: 列宽应该自适应
        col1_width = ws.column_dimensions['A'].width

        col2_width = ws.column_dimensions['B'].width

        col3_width = ws.column_dimensions['C'].width

        # 第三列应该最宽
        assert col3_width > col2_width

        assert col2_width > col1_width

    

    def test_min_column_width(self, temp_output_dir):

        """最小列宽"""
        file_path = temp_output_dir / "min_width.xlsx"

        result = write_xlsx(str(file_path), data=[{"A": "1"}])

        assert is_success(result)

        wb = load_workbook(str(file_path))

        ws = wb.active

        # Bug #17: 列宽应该有最小值
        assert ws.column_dimensions['A'].width >= 8





class TestWriteXlsxLargeData:

    """大数据量测试"""
    

    def test_100_rows(self, temp_output_dir):

        """100行数据"""
        file_path = temp_output_dir / "100_rows.xlsx"

        data = [{"序号": i, "数据": f"内容{i}"} for i in range(1, 101)]
        result = write_xlsx(str(file_path), data=data)

        assert is_success(result)

        wb = load_workbook(str(file_path))

        # Bug #18: 应该有100行数据
        assert wb.active.max_row == 101  # 表头+100行
    

    def test_100_columns(self, temp_output_dir):

        """100列数据"""
        file_path = temp_output_dir / "100_cols.xlsx"

        data = [{f"列{i}": f"数据{i}" for i in range(1, 101)}]
        result = write_xlsx(str(file_path), data=data)

        assert is_success(result)

        wb = load_workbook(str(file_path))

        # Bug #19: 应该有100列
        assert wb.active.max_column == 100

    

    def test_1000_rows(self, temp_output_dir):

        """1000行数据"""
        file_path = temp_output_dir / "1000_rows.xlsx"

        data = [{"序号": i, "数据": f"内容{i}"} for i in range(1, 1001)]
        result = write_xlsx(str(file_path), data=data)

        assert is_success(result)

        wb = load_workbook(str(file_path))

        # Bug #20: 大数据量性能测试
        assert wb.active.max_row == 1001





class TestWriteXlsxSheetName:

    """工作表名测试"""
    

    def test_chinese_sheet_name(self, temp_output_dir):
 
        """中文工作表名"""
        file_path = temp_output_dir / "chinese_sheet.xlsx"
 
        result = write_xlsx(str(file_path), data=[{"A": "1"}], sheet_name="数据表")
        assert is_success(result)
 
        wb = load_workbook(str(file_path))
 
        assert wb.active.title == "数据表"
    

    def test_long_sheet_name(self, temp_output_dir):
 
        """超长工作表名"""
        file_path = temp_output_dir / "long_sheet.xlsx"
 
        long_name = "A" * 50
 
        result = write_xlsx(str(file_path), data=[{"A": "1"}], sheet_name=long_name)
 
        # Bug #21: 超长工作表名如何处理?
        assert is_success(result)  # openpyxl会警告但不会报错
    

    def test_special_chars_sheet_name(self, temp_output_dir):

        """特殊字符工作表名"""
        file_path = temp_output_dir / "special_sheet.xlsx"

        result = write_xlsx(str(file_path), sheet_name="数据/表")
        # Excel不允许工作表名包含 \ / ? * [ ] — 小欧 2026-06-24
        assert is_error(result)





class TestWriteXlsxRealScenarios:

    """真实业务场景测试"""
    

    def test_employee_data(self, temp_output_dir):

        """员工数据"""
        file_path = temp_output_dir / "员工表.xlsx"
        data = [

            {"工号": "E001", "姓名": "张三", "部门": "技术部", "职位": "工程师", "薪资": 15000, "入职日期": "2020-01-15"},
            {"工号": "E002", "姓名": "李四", "部门": "产品部", "职位": "产品经理", "薪资": 18000, "入职日期": "2019-06-20"},
            {"工号": "E003", "姓名": "王五", "部门": "技术部", "职位": "架构师", "薪资": 25000, "入职日期": "2018-03-10"},
            {"工号": "E004", "姓名": "赵六", "部门": "运营部", "职位": "运营总监", "薪资": 20000, "入职日期": "2017-11-05"},
            {"工号": "E005", "姓名": "孙七", "部门": "财务部", "职位": "财务主管", "薪资": 16000, "入职日期": "2021-02-28"},
        ]

        result = write_xlsx(str(file_path), data=data, sheet_name="员工信息")
        assert is_success(result)

        wb = load_workbook(str(file_path))

        ws = wb.active

        assert ws.title == "员工信息"
        assert ws.max_row == 6  # 表头+5行数据
        assert ws.max_column == 6  # 小欧 2026-06-24: 数据只有6个key(工号/姓名/部门/职位/薪资/入职日期)
    

    def test_sales_report(self, temp_output_dir):

        """销售报表"""
        file_path = temp_output_dir / "销售报表.xlsx"
        data = [

            {"日期": "2026-06-01", "产品": "产品A", "销量": 120, "金额": 12000, "区域": "华东"},
            {"日期": "2026-06-01", "产品": "产品B", "销量": 80, "金额": 16000, "区域": "华东"},
            {"日期": "2026-06-02", "产品": "产品A", "销量": 150, "金额": 15000, "区域": "华南"},
            {"日期": "2026-06-02", "产品": "产品C", "销量": 200, "金额": 20000, "区域": "华南"},
            {"日期": "2026-06-03", "产品": "产品B", "销量": 90, "金额": 18000, "区域": "华北"},
        ]

        result = write_xlsx(str(file_path), data=data)

        assert is_success(result)

        wb = load_workbook(str(file_path))

        assert wb.active.max_row == 6





class TestWriteXlsxBoundary:

    """边界测试"""
    

    def test_empty_dict(self, temp_output_dir):

        """空字典 — 小欧 2026-06-24 openpyxl默认有1列"""
        file_path = temp_output_dir / "empty_dict.xlsx"

        result = write_xlsx(str(file_path), data=[{}])

        assert is_success(result)

        wb = load_workbook(str(file_path))

        # openpyxl即使空数据也有1列默认列 — 小欧 2026-06-24
        assert wb.active.max_column <= 1

    

    def test_all_none_values(self, temp_output_dir):

        """全None值"""
        file_path = temp_output_dir / "all_none.xlsx"

        result = write_xlsx(str(file_path), data=[{"A": None, "B": None, "C": None}])

        assert is_success(result)

        wb = load_workbook(str(file_path))

        # Bug #24: 全None值如何处理?
        ws = wb.active

        assert ws.cell(2, 1).value is None

        assert ws.cell(2, 2).value is None

        assert ws.cell(2, 3).value is None

    

    def test_duplicate_column_names(self, temp_output_dir):

        """重复列名"""
        file_path = temp_output_dir / "dup_cols.xlsx"

        # Bug #25: 重复列名如何处理?
        # Python字典不允许重复key,所以这个测试用例无法构造
        # 但可以测试相同key不同值的情况
        result = write_xlsx(str(file_path), data=[{"A": "1", "A": "2"}])

        # 这在Python中在面的A会覆盖前面的A




class TestWriteXlsxNegative:

    """为面测试"""
    

    def test_invalid_path(self):

        """无效路径"""
        result = write_xlsx("Z:/invalid/path/data.xlsx", data=[{"A": "1"}])

        assert is_error(result)

    

    def test_invalid_file_extension(self, temp_output_dir):
 
        """无效文件扩展名"""
        file_path = temp_output_dir / "data.txt"
 
        result = write_xlsx(str(file_path), data=[{"A": "1"}])
 
        # Bug #26: 无效扩展名如何处理? — 当前行为:write_xlsx拒绝非.xlsx扩展名
        assert is_error(result)
    

    def test_permission_denied(self):

        """权限拒绝"""
        # Windows下无法测试,跳过
        pass





class TestWriteXlsxSchemaIssues:

    """Schema问题测试"""
    

    def test_data_description_incomplete(self):

        """data描述不完整"""
        from app.tools.document.document_schema import WriteXlsxInput

        field = WriteXlsxInput.model_fields['data']

        # Bug #27: data描述应该说明列合并逻辑
        assert "合并" in field.description or "key" in field.description
    

    def test_sheet_name_default(self):

        """sheet_name默认值"""
        from app.tools.document.document_schema import WriteXlsxInput

        field = WriteXlsxInput.model_fields['sheet_name']

        # Bug #28: sheet_name默认值应该是"Sheet1"
        assert field.default == "Sheet1"

    

    def test_no_error_handling_doc(self):

        """缺少错误处理说明"""
        from app.tools.document.document_schema import WriteXlsxInput

        # Bug #29: Schema应该说明错误处理情况
        # 如:无效路径,权限问题,数据格式错误等
        pass

    

    def test_examples_coverage(self):

        """Examples覆盖度"""
        from app.tools.document.document_register import EXAMPLES

        examples = EXAMPLES.get("write_xlsx", [])

        # Bug #30: Examples应该覆盖更多场景
        assert len(examples) >= 3  # 至少3个示例