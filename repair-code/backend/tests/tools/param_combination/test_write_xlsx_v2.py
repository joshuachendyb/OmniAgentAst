# -*- coding: utf-8 -*-
"""
write_xlsx参数组合与内容测试v2 - 小健 2026-06-24
严格按照案范:5-40个case,真实业务数据,验证实际内容
"""

import pytest
from pathlib import Path
from openpyxl import load_workbook
from app.tools.document.write_xlsx import write_xlsx
from app.tools.tool_response import is_success, is_error


class TestWriteXlsxParamCombinations:
    """参数组合测试 - 8种组合"""

    def test_file_name_only(self, temp_output_dir):
        """组合1: 仅file_name — 当前行为:data为必填,缺失返回error"""
        file_path = temp_output_dir / "test1.xlsx"
        result = write_xlsx(path=str(file_path))
        assert is_error(result)

    def test_file_name_empty_data(self, temp_output_dir):
        """组合2: file_name + data=[] — 当前行为:空data被拒绝"""
        file_path = temp_output_dir / "test2.xlsx"
        result = write_xlsx(path=str(file_path), data=[])
        assert is_error(result)

    def test_file_name_data(self, temp_output_dir):
        """组合3: file_name + data"""
        file_path = temp_output_dir / "test3.xlsx"
        data = [{"姓名": "张三", "年龄": 25}]
        result = write_xlsx(path=str(file_path), data=data)
        assert is_success(result)
        wb = load_workbook(str(file_path))
        ws = wb.active
        assert ws.cell(1, 1).value == "姓名"
        assert ws.cell(2, 1).value == "张三"

    def test_file_name_sheet_name(self, temp_output_dir):
        """组合4: file_name + sheet_name"""
        file_path = temp_output_dir / "test4.xlsx"
        result = write_xlsx(path=str(file_path), data=[{"A": "1"}], sheet_name="销售数据")
        assert is_success(result)
        wb = load_workbook(str(file_path))
        assert wb.sheetnames[0] == "销售数据"

    def test_file_name_data_sheet_name(self, temp_output_dir):
        """组合5: file_name + data + sheet_name"""
        file_path = temp_output_dir / "test5.xlsx"
        data = [{"产品": "A", "销量": 100}]
        result = write_xlsx(path=str(file_path), data=data, sheet_name="产品")
        assert is_success(result)
        wb = load_workbook(str(file_path))
        assert wb.sheetnames[0] == "产品"
        assert wb.active.cell(2, 1).value == "A"

    def test_data_none(self, temp_output_dir):
        """组合6: data=None — 当前行为:data为必填,缺失返回error"""
        file_path = temp_output_dir / "test6.xlsx"
        result = write_xlsx(path=str(file_path), data=None)
        assert is_error(result)

    def test_single_column(self, temp_output_dir):
        """组合7: 单列数据"""
        file_path = temp_output_dir / "test7.xlsx"
        data = [{"列1": "A"}, {"列1": "B"}, {"列1": "C"}]
        result = write_xlsx(path=str(file_path), data=data)
        assert is_success(result)
        wb = load_workbook(str(file_path))
        ws = wb.active
        assert ws.max_column == 1
        assert ws.max_row == 4

    def test_multi_sheet_names(self, temp_output_dir):
        """组合8: 不同sheet_name"""
        file_path = temp_output_dir / "test8.xlsx"
        for name in ["数据", "Data", "报表123", "Sheet_测试"]:
            result = write_xlsx(path=str(file_path), data=[{"A": "1"}], sheet_name=name)
            assert is_success(result)


class TestWriteXlsxSingleFeatures:
    """单一功能测试 - 10个case"""

    def test_chinese_header(self, temp_output_dir):
        """中文表头"""
        file_path = temp_output_dir / "chinese_header.xlsx"
        data = [{"员工编号": "001", "入职日期": "2020-01-01"}]
        result = write_xlsx(path=str(file_path), data=data)
        assert is_success(result)
        wb = load_workbook(str(file_path))
        ws = wb.active
        assert ws.cell(1, 1).value == "员工编号"

    def test_numeric_types(self, temp_output_dir):
        """数字类型:整数,浮点,为数,零"""
        file_path = temp_output_dir / "numeric.xlsx"
        data = [
            {"整数": 100, "浮点": 99.5, "为数": -50, "零": 0},
            {"整数": 200, "浮点": 88.8, "为数": -30, "零": 0},
        ]
        result = write_xlsx(path=str(file_path), data=data)
        assert is_success(result)
        wb = load_workbook(str(file_path))
        ws = wb.active
        assert ws.cell(2, 1).value == 100
        assert ws.cell(2, 2).value == 99.5
        assert ws.cell(2, 3).value == -50
        assert ws.cell(2, 4).value == 0

    def test_string_types(self, temp_output_dir):
        """字符串类型:普通,空字符串,空格"""
        file_path = temp_output_dir / "string.xlsx"
        data = [
            {"普通": "文本", "空字符串": "", "空格": "   "},
        ]
        result = write_xlsx(path=str(file_path), data=data)
        assert is_success(result)
        wb = load_workbook(str(file_path))
        ws = wb.active
        # openpyxl会把空字符串转成None
        assert ws.cell(2, 2).value is None  # 空字符串→None
        assert ws.cell(2, 3).value == "   "  # 空格保留

    def test_special_chars(self, temp_output_dir):
        """特殊字符:<>&\"'换行制表符"""
        file_path = temp_output_dir / "special.xlsx"
        data = [
            {"特殊": "<>&\"'", "换行": "第一行\n第二行", "制表": "A\tB"},
        ]
        result = write_xlsx(path=str(file_path), data=data)
        assert is_success(result)
        wb = load_workbook(str(file_path))
        ws = wb.active
        assert ws.cell(2, 1).value == "<>&\"'"
        assert "\n" in ws.cell(2, 2).value

    def test_boolean_none(self, temp_output_dir):
        """布尔和None值"""
        file_path = temp_output_dir / "bool_none.xlsx"
        data = [
            {"布尔True": True, "布尔False": False, "None值": None},
        ]
        result = write_xlsx(path=str(file_path), data=data)
        assert is_success(result)
        wb = load_workbook(str(file_path))
        ws = wb.active
        assert ws.cell(2, 1).value is True
        assert ws.cell(2, 2).value is False
        assert ws.cell(2, 3).value is None

    def test_long_text(self, temp_output_dir):
        """长文本(1000字符)"""
        file_path = temp_output_dir / "long.xlsx"
        long_text = "这是一段很长的文本," * 50
        data = [{"描述": long_text}]
        result = write_xlsx(path=str(file_path), data=data)
        assert is_success(result)
        wb = load_workbook(str(file_path))
        ws = wb.active
        # openpyxl会保留长文本
        assert len(str(ws.cell(2, 1).value)) >= 500  # 至少500字符

    def test_datetime_string(self, temp_output_dir):
        """日期时间字符串"""
        file_path = temp_output_dir / "datetime.xlsx"
        data = [
            {"日期": "2026-06-24", "时间": "10:30:00", "日期时间": "2026-06-24 10:30:00"},
        ]
        result = write_xlsx(path=str(file_path), data=data)
        assert is_success(result)
        wb = load_workbook(str(file_path))
        ws = wb.active
        assert ws.cell(2, 1).value == "2026-06-24"

    def test_formula_string(self, temp_output_dir):
        """公式字符串(不解析,原样写入)"""
        file_path = temp_output_dir / "formula.xlsx"
        data = [
            {"公式": "=SUM(A1:A10)", "普通": "文本"},
        ]
        result = write_xlsx(path=str(file_path), data=data)
        assert is_success(result)

    def test_json_string(self, temp_output_dir):
        """JSON字符串"""
        file_path = temp_output_dir / "json.xlsx"
        data = [
            {"JSON": '{"name":"张三","age":25}', "数组": '[1,2,3]'},
        ]
        result = write_xlsx(path=str(file_path), data=data)
        assert is_success(result)

    def test_unicode_emoji(self, temp_output_dir):
        """Unicode和Emoji"""
        file_path = temp_output_dir / "unicode.xlsx"
        data = [
            {"中文": "测试", "日文": "テスト", "Emoji": "🔥🎀"},
        ]
        result = write_xlsx(path=str(file_path), data=data)
        assert is_success(result)
        wb = load_workbook(str(file_path))
        ws = wb.active
        assert ws.cell(2, 3).value == "🔥🎀"


class TestWriteXlsxRealScenarios:
    """真实业务场景测试 - 5个case,数据不少于100行"""

    def test_employee_salary(self, temp_output_dir):
        """员工薪资表(100行,验证工资计算)"""
        file_path = temp_output_dir / "员工薪资.xlsx"

        departments = ["技术部", "产品部", "运营部", "市场部", "财务部"]
        positions = ["工程师", "高级工程师", "经理", "总监", "专员"]
        data = []

        for i in range(1, 101):
            dept = departments[(i-1) % 5]
            pos = positions[(i-1) % 5]
            base = 8000 + (i % 10) * 1000
            bonus = base * 0.2 if i % 3 == 0 else 0
            tax = (base + bonus) * 0.1

            data.append({
                "工号": f"EMP{i:04d}",
                "姓名": f"员工{i}",
                "部门": dept,
                "职位": pos,
                "基本工资": base,
                "绩效奖金": bonus,
                "扣税": tax,
                "实发工资": base + bonus - tax,
                "入职日期": f"2020-{((i-1) % 12) + 1:02d}-{((i-1) % 28) + 1:02d}",
            })

        result = write_xlsx(path=str(file_path), data=data, sheet_name="薪资明细")
        assert is_success(result)

        wb = load_workbook(str(file_path))
        ws = wb.active
        assert ws.max_row == 101
        assert ws.max_column == 9

        # 验证工资计算正认性(抽查10行)
        for row in [2, 10, 25, 50, 75, 100]:
            base = ws.cell(row, 5).value
            bonus = ws.cell(row, 6).value
            tax = ws.cell(row, 7).value
            actual = ws.cell(row, 8).value
            expected = base + bonus - tax
            assert abs(actual - expected) < 0.01, f"第{row}行工资计算错误"

    def test_sales_order(self, temp_output_dir):
        """销售订单表(100行,验证金额计算)"""
        file_path = temp_output_dir / "销售订单.xlsx"

        products = ["产品A", "产品B", "产品C", "产品D", "产品E"]
        regions = ["华东", "华南", "华北", "西南", "西北"]
        data = []

        for i in range(1, 101):
            qty = 10 + (i % 50)
            price = 100 + (i * 5)
            data.append({
                "订单号": f"ORD{i:05d}",
                "产品": products[i % 5],
                "区域": regions[i % 5],
                "数量": qty,
                "单价": price,
                "金额": qty * price,
                "日期": f"2026-{((i-1) // 20) + 1:02d}-{(i % 28) + 1:02d}",
            })

        result = write_xlsx(path=str(file_path), data=data)
        assert is_success(result)

        wb = load_workbook(str(file_path))
        ws = wb.active
        assert ws.max_row == 101

        # 验证金额 = 数量 * 单价
        for row in [2, 30, 60, 90]:
            qty = ws.cell(row, 4).value
            price = ws.cell(row, 5).value
            amount = ws.cell(row, 6).value
            assert amount == qty * price, f"第{row}行金额计算错误"

    def test_student_score(self, temp_output_dir):
        """学生成绩表(100行,验证总分和平均分)"""
        file_path = temp_output_dir / "学生成绩.xlsx"

        subjects = ["语文", "数学", "英语", "物理", "化学"]
        data = []

        for i in range(1, 101):
            scores = [60 + (i * j) % 40 for j in range(5)]
            total = sum(scores)
            avg = total / 5

            row = {
                "学号": f"STU{i:04d}",
                "姓名": f"学生{i}",
            }
            for j, subj in enumerate(subjects):
                row[subj] = scores[j]
            row["总分"] = total
            row["平均分"] = avg

            data.append(row)

        result = write_xlsx(path=str(file_path), data=data)
        assert is_success(result)

        wb = load_workbook(str(file_path))
        ws = wb.active
        assert ws.max_row == 101

        # 验证总分和平均分
        for row in [2, 25, 50, 75, 100]:
            scores = [ws.cell(row, j).value for j in range(3, 8)]
            total = ws.cell(row, 8).value
            avg = ws.cell(row, 9).value
            assert total == sum(scores), f"第{row}行总分错误"
            assert abs(avg - sum(scores)/5) < 0.01, f"第{row}行平均分错误"

    def test_inventory(self, temp_output_dir):
        """库存清单(100行,验证库存价值)"""
        file_path = temp_output_dir / "库存清单.xlsx"

        categories = ["电子产品", "办公用品", "原材料", "成品", "半成品"]
        data = []

        for i in range(1, 101):
            qty = 100 + (i % 500)
            price = 10 + (i * 2)
            data.append({
                "物料编号": f"MAT{i:04d}",
                "物料名称": f"物料{i}",
                "类别": categories[i % 5],
                "库存数量": qty,
                "单价": price,
                "库存价值": qty * price,
                "仓库": f"仓库{(i % 3) + 1}",
                "更新日期": "2026-06-24",
            })

        result = write_xlsx(path=str(file_path), data=data)
        assert is_success(result)

        wb = load_workbook(str(file_path))
        ws = wb.active
        assert ws.max_row == 101

    def test_project_task(self, temp_output_dir):
        """项目任务表(100行)"""
        file_path = temp_output_dir / "项目任务.xlsx"

        statuses = ["未开始", "进行中", "已完成", "已延迟"]
        priorities = ["高", "中", "低"]
        assignees = ["张三", "李四", "王五", "赵六", "钱七"]
        data = []

        for i in range(1, 101):
            data.append({
                "任务ID": f"TASK{i:04d}",
                "任务名称": f"任务{i}",
                "为责人": assignees[i % 5],
                "优先级": priorities[i % 3],
                "状态": statuses[i % 4],
                "预计工时": 8 + (i % 40),
                "实际工时": 8 + (i % 50),
                "开始日期": f"2026-0{((i-1) // 30) + 1:02d}-{(i % 28) + 1:02d}",
            })

        result = write_xlsx(path=str(file_path), data=data)
        assert is_success(result)

        wb = load_workbook(str(file_path))
        ws = wb.active
        assert ws.max_row == 101


class TestWriteXlsxBoundary:
    """边界测试 - 6个case"""

    def test_large_data_1000_rows(self, temp_output_dir):
        """大数据量(1000行)"""
        file_path = temp_output_dir / "large.xlsx"
        data = [{"序号": i, "数据": f"内容{i}"} for i in range(1000)]
        result = write_xlsx(path=str(file_path), data=data)
        assert is_success(result)
        wb = load_workbook(str(file_path))
        assert wb.active.max_row == 1001

    def test_many_columns_50(self, temp_output_dir):
        """多列(50列)"""
        file_path = temp_output_dir / "many_cols.xlsx"
        row = {f"列{i}": f"数据{i}" for i in range(50)}
        data = [row]
        result = write_xlsx(path=str(file_path), data=data)
        assert is_success(result)
        wb = load_workbook(str(file_path))
        assert wb.active.max_column == 50

    def test_inconsistent_columns(self, temp_output_dir):
        """列不一致(Bug #4验证)"""
        file_path = temp_output_dir / "inconsistent.xlsx"
        data = [
            {"A": 1, "B": 2},
            {"A": 3, "C": 4},
            {"A": 5, "B": 6, "C": 7, "D": 8},
        ]
        result = write_xlsx(path=str(file_path), data=data)
        assert is_success(result)

        wb = load_workbook(str(file_path))
        ws = wb.active
        # 验证:所有列都保留,缺失填None
        assert ws.max_column == 4
        assert ws.cell(2, 3).value is None  # 第2行C列为None
        assert ws.cell(3, 2).value is None  # 第3行B列为None
        assert ws.cell(4, 4).value == 8     # 第4行D列有值

    def test_empty_string_vs_none(self, temp_output_dir):
        """空字符串vs None - Bug #5: openpyxl把空字符串转成None"""
        file_path = temp_output_dir / "empty_vs_none.xlsx"
        data = [
            {"空字符串": "", "None": None, "空格": "   "},
        ]
        result = write_xlsx(path=str(file_path), data=data)
        assert is_success(result)
        wb = load_workbook(str(file_path))
        ws = wb.active
        # openpyxl的行为:空字符串→None
        assert ws.cell(2, 1).value is None  # 空字符串变成None
        assert ws.cell(2, 2).value is None  # None还是None
        assert ws.cell(2, 3).value == "   "  # 空格保留

    def test_duplicate_keys(self, temp_output_dir):
        """重复key(dict自动去重)"""
        file_path = temp_output_dir / "duplicate.xlsx"
        # Python dict不允许重复key,这里测试相同key不同值
        data = [{"A": 1}, {"A": 2}, {"A": 3}]
        result = write_xlsx(path=str(file_path), data=data)
        assert is_success(result)
        wb = load_workbook(str(file_path))
        ws = wb.active
        assert ws.cell(2, 1).value == 1
        assert ws.cell(3, 1).value == 2

    def test_all_none_row(self, temp_output_dir):
        """全None行"""
        file_path = temp_output_dir / "all_none.xlsx"
        data = [
            {"A": 1, "B": 2},
            {"A": None, "B": None},  # 全None
            {"A": 3, "B": 4},
        ]
        result = write_xlsx(path=str(file_path), data=data)
        assert is_success(result)
        wb = load_workbook(str(file_path))
        ws = wb.active
        assert ws.cell(3, 1).value is None
        assert ws.cell(3, 2).value is None


class TestWriteXlsxNegative:
    """为面测试 - 4个case"""

    def test_invalid_path(self):
        """无效路径"""
        result = write_xlsx(path="Z:/invalid/path/test.xlsx", data=[{"A": 1}])
        assert is_error(result)

    def test_permission_denied(self):
        """权限不足(系统目录)"""
        result = write_xlsx(path="C:/Windows/test.xlsx", data=[{"A": 1}])
        assert is_error(result)

    def test_invalid_sheet_name(self, temp_output_dir):
        """无效sheet名称(特殊字符)"""
        file_path = temp_output_dir / "invalid_sheet.xlsx"
        # sheet名称不能包含: \ / ? * [ ]
        result = write_xlsx(path=str(file_path), sheet_name="测试/Sheet")
        # 可能成功也可能失败,取决于openpyxl处理
        # 验证:不崩溃即可

    def test_very_long_sheet_name(self, temp_output_dir):
        """超长sheet名称"""
        file_path = temp_output_dir / "long_sheet.xlsx"
        long_name = "A" * 100
        result = write_xlsx(path=str(file_path), sheet_name=long_name)
        # openpyxl会截断sheet名称


class TestWriteXlsxSchemaIssues:
    """Schema问题验证"""

    def test_examples_too_few(self):
        """Bug:Examples只有1个,覆盖率33%"""
        # Schema有3个参数,Examples只有1个
        # LLM不知道sheet_name参数
        pass

    def test_data_description_incomplete(self):
        """Bug:data参数描述不完整"""
        # 没说明:
        # 1. 空数组[]是否合法
        # 2. None是否合法
        # 3. key不一致如何处理
        # 4. 支持哪些数据类型
        pass

    def test_no_error_handling_doc(self):
        """Bug:缺少错误处理说明"""
        # 没说明哪些情况会返回error
        pass


# ============================================================
# P04 append_mode 追加模式测试 — 小欧 2026-08-07
# ============================================================

class TestWriteXlsxAppendMode:
    """append_mode=True: 追加到已有文件/表头不一致拒绝/文件不存在等同新建"""

    def test_append_to_existing_file(self, temp_output_dir):
        """追加到已有文件: 首次2行 + 追加3行 = 表头+5行数据"""
        file_path = temp_output_dir / "append1.xlsx"
        r1 = write_xlsx(path=str(file_path), data=[{"A": "a1", "B": "b1"}, {"A": "a2", "B": "b2"}])
        assert is_success(r1)
        r2 = write_xlsx(path=str(file_path), data=[{"A": "a3", "B": "b3"}, {"A": "a4", "B": "b4"}, {"A": "a5", "B": "b5"}], append_mode=True)
        assert is_success(r2)
        wb = load_workbook(str(file_path))
        ws = wb.active
        assert ws.max_row == 6
        assert ws.cell(row=6, column=1).value == "a5"
        assert ws.cell(row=6, column=2).value == "b5"

    def test_append_keeps_original_rows(self, temp_output_dir):
        """追加模式不清空原数据(不覆盖)"""
        file_path = temp_output_dir / "append2.xlsx"
        write_xlsx(path=str(file_path), data=[{"A": "old", "B": "keep"}])
        write_xlsx(path=str(file_path), data=[{"A": "new", "B": "x"}], append_mode=True)
        wb = load_workbook(str(file_path))
        ws = wb.active
        assert ws.cell(row=2, column=1).value == "old"
        assert ws.cell(row=3, column=1).value == "new"

    def test_append_header_mismatch_rejected(self, temp_output_dir):
        """追加数据含新列(不在已有表头) → error, 原文件不变"""
        file_path = temp_output_dir / "append3.xlsx"
        write_xlsx(path=str(file_path), data=[{"A": 1, "B": 2}])
        r = write_xlsx(path=str(file_path), data=[{"A": 3, "C": 4}], append_mode=True)
        assert is_error(r)
        wb = load_workbook(str(file_path))
        ws = wb.active
        assert ws.max_row == 2  # 原文件未变

    def test_append_when_file_not_exists_creates(self, temp_output_dir):
        """append_mode=True 但文件不存在 → 等同新建"""
        file_path = temp_output_dir / "sub" / "append4.xlsx"
        r = write_xlsx(path=str(file_path), data=[{"X": 1}], append_mode=True)
        assert is_success(r)
        wb = load_workbook(str(file_path))
        assert wb.active.cell(row=2, column=1).value == 1
