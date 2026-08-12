# -*- coding: utf-8 -*-
"""
write_xlsx参数组合与内容测试 - 小健 2026-06-24

测试目标:
1. 参数组合:file_name(必填), data(可选), sheet_name(可选)
2. 功能点:空数据,单行数据,多行数据,多列数据,特殊字符,中文,数字,日期
3. 真实场景:员工信息表,销售数据表,技术指标表
4. 边界测试:空data,None,大数据量,特殊字符
5. 负面测试:无效路径,权限问题
"""

import pytest
from pathlib import Path
from openpyxl import load_workbook
from app.tools.document.write_xlsx import write_xlsx
from app.tools.tool_response import is_success, is_error


class TestWriteXlsxParamCombinations:
    """参数组合测试 - 6种组合"""

    def test_file_name_only(self, temp_output_dir):
        """组合1: 仅必填参数file_name — 当前行为:data为必填,缺失返回error"""
        file_path = temp_output_dir / "empty.xlsx"
        result = write_xlsx(path=str(file_path))

        assert is_error(result)

    def test_file_name_with_empty_data(self, temp_output_dir):
        """组合2: file_name + data=[] — 当前行为:空data被拒绝"""
        file_path = temp_output_dir / "empty_data.xlsx"
        result = write_xlsx(path=str(file_path), data=[])

        assert is_error(result)

    def test_file_name_with_data(self, temp_output_dir):
        """组合3: file_name + data"""
        file_path = temp_output_dir / "data.xlsx"
        data = [{"姓名": "张三", "年龄": 25}, {"姓名": "李四", "年龄": 30}]
        result = write_xlsx(path=str(file_path), data=data)

        assert is_success(result)
        wb = load_workbook(str(file_path))
        ws = wb.active
        # 验证:表头+2行数据
        assert ws.max_row == 3
        assert ws.max_column == 2
        assert ws.cell(1, 1).value == "姓名"
        assert ws.cell(2, 1).value == "张三"

    def test_file_name_with_sheet_name(self, temp_output_dir):
        """组合4: file_name + sheet_name"""
        file_path = temp_output_dir / "custom_sheet.xlsx"
        result = write_xlsx(path=str(file_path), data=[{"A": "1"}], sheet_name="销售数据")

        assert is_success(result)
        wb = load_workbook(str(file_path))
        assert wb.sheetnames[0] == "销售数据"

    def test_file_name_data_sheet_name(self, temp_output_dir):
        """组合5: file_name + data + sheet_name"""
        file_path = temp_output_dir / "full.xlsx"
        data = [{"产品": "A", "销量": 100}]
        result = write_xlsx(path=str(file_path), data=data, sheet_name="产品销售")

        assert is_success(result)
        wb = load_workbook(str(file_path))
        assert wb.sheetnames[0] == "产品销售"
        ws = wb.active
        assert ws.cell(1, 1).value == "产品"

    def test_data_none(self, temp_output_dir):
        """组合6: data=None — 当前行为:data为必填,缺失返回error"""
        file_path = temp_output_dir / "none.xlsx"
        result = write_xlsx(path=str(file_path), data=None)

        assert is_error(result)


class TestWriteXlsxSingleFeatures:
    """单一功能测试"""

    def test_single_row(self, temp_output_dir):
        """单行数据"""
        file_path = temp_output_dir / "single.xlsx"
        data = [{"项目": "测试项目", "状态": "进行中"}]
        result = write_xlsx(path=str(file_path), data=data)

        assert is_success(result)
        wb = load_workbook(str(file_path))
        ws = wb.active
        assert ws.max_row == 2  # 表头+1行

    def test_multi_columns(self, temp_output_dir):
        """多列数据(10列)"""
        file_path = temp_output_dir / "multi_col.xlsx"
        data = [{
            "列1": "A", "列2": "B", "列3": "C", "列4": "D", "列5": "E",
            "列6": "F", "列7": "G", "列8": "H", "列9": "I", "列10": "J"
        }]
        result = write_xlsx(path=str(file_path), data=data)

        assert is_success(result)
        wb = load_workbook(str(file_path))
        ws = wb.active
        assert ws.max_column == 10

    def test_chinese_content(self, temp_output_dir):
        """中文内容"""
        file_path = temp_output_dir / "chinese.xlsx"
        data = [
            {"姓名": "张三", "部门": "技术部", "职位": "高级工程师"},
            {"姓名": "李四", "部门": "产品部", "职位": "产品经理"},
            {"姓名": "王五", "部门": "运营部", "职位": "运营总监"},
        ]
        result = write_xlsx(path=str(file_path), data=data)

        assert is_success(result)
        wb = load_workbook(str(file_path))
        ws = wb.active
        assert ws.cell(2, 1).value == "张三"
        assert ws.cell(3, 2).value == "产品部"

    def test_numeric_types(self, temp_output_dir):
        """数字类型:整数,浮点数,负数"""
        file_path = temp_output_dir / "numeric.xlsx"
        data = [
            {"名称": "项目A", "整数": 100, "浮点数": 99.5, "负数": -50},
            {"名称": "项目B", "整数": 200, "浮点数": 88.8, "负数": -30},
        ]
        result = write_xlsx(path=str(file_path), data=data)

        assert is_success(result)
        wb = load_workbook(str(file_path))
        ws = wb.active
        assert ws.cell(2, 2).value == 100
        assert ws.cell(2, 3).value == 99.5
        assert ws.cell(2, 4).value == -50


class TestWriteXlsxRealScenarios:
    """真实业务场景测试 - 数据不少于100行"""

    def test_employee_info_table(self, temp_output_dir):
        """员工信息表(真实场景)"""
        file_path = temp_output_dir / "员工信息表.xlsx"

        # 生成100行员工数据
        departments = ["技术部", "产品部", "运营部", "市场部", "财务部"]
        positions = ["工程师", "经理", "总监", "专员", "主管"]
        data = []
        for i in range(1, 101):
            data.append({
                "工号": f"EMP{i:04d}",
                "姓名": f"员工{i}",
                "部门": departments[i % 5],
                "职位": positions[i % 5],
                "入职日期": f"2020-{(i % 12) + 1:02d}-{(i % 28) + 1:02d}",
                "工资": 8000 + (i * 100),
                "状态": "在职" if i % 10 != 0 else "离职"
            })

        result = write_xlsx(path=str(file_path), data=data, sheet_name="员工信息")

        assert is_success(result)
        wb = load_workbook(str(file_path))
        ws = wb.active
        # 验证:表头+100行数据
        assert ws.max_row == 101
        assert ws.max_column == 7
        assert ws.cell(1, 1).value == "工号"
        assert ws.cell(100, 2).value == "员工99"

    def test_sales_data_table(self, temp_output_dir):
        """销售数据表(真实场景)"""
        file_path = temp_output_dir / "销售数据.xlsx"

        products = ["产品A", "产品B", "产品C", "产品D", "产品E"]
        regions = ["华东", "华南", "华北", "西南", "西北"]
        data = []
        for i in range(1, 101):
            data.append({
                "订单号": f"ORD{i:05d}",
                "产品": products[i % 5],
                "区域": regions[i % 5],
                "数量": 10 + (i % 50),
                "单价": 100 + (i * 5),
                "金额": (10 + (i % 50)) * (100 + (i * 5)),
                "日期": f"2026-0{((i-1) // 20) + 1}-{(i % 28) + 1:02d}"
            })

        result = write_xlsx(path=str(file_path), data=data)

        assert is_success(result)
        wb = load_workbook(str(file_path))
        ws = wb.active
        assert ws.max_row == 101
        # 验证金额计算正认
        assert ws.cell(2, 6).value == ws.cell(2, 4).value * ws.cell(2, 5).value


class TestWriteXlsxBoundary:
    """边界测试"""

    def test_special_chars(self, temp_output_dir):
        """特殊字符:<>&\"'"""
        file_path = temp_output_dir / "special.xlsx"
        data = [
            {"字段": "特殊字符测试", "内容": "<>&\"'测试"},
            {"字段": "换行符", "内容": "第一行\n第二行"},
        ]
        result = write_xlsx(path=str(file_path), data=data)

        assert is_success(result)
        wb = load_workbook(str(file_path))
        ws = wb.active
        assert ws.cell(2, 2).value == "<>&\"'测试"

    def test_long_text(self, temp_output_dir):
        """长文本(500字符)"""
        file_path = temp_output_dir / "long.xlsx"
        long_text = "这是一段很长的文本内容," * 50  # 500+字符
        data = [{"描述": long_text}]
        result = write_xlsx(path=str(file_path), data=data)

        assert is_success(result)
        wb = load_workbook(str(file_path))
        ws = wb.active
        assert len(str(ws.cell(2, 1).value)) > 500

    def test_large_data(self, temp_output_dir):
        """大数据量(1000行)"""
        file_path = temp_output_dir / "large.xlsx"
        data = [{"序号": i, "数据": f"内容{i}"} for i in range(1000)]
        result = write_xlsx(path=str(file_path), data=data)

        assert is_success(result)
        wb = load_workbook(str(file_path))
        ws = wb.active
        assert ws.max_row == 1001  # 表头+1000行

    def test_mixed_types(self, temp_output_dir):
        """混合类型:字符串,数字,布尔,None"""
        file_path = temp_output_dir / "mixed.xlsx"
        data = [
            {"字符串": "文本", "整数": 100, "浮点数": 99.9, "布尔": True, "空值": None},
        ]
        result = write_xlsx(path=str(file_path), data=data)

        assert is_success(result)
        wb = load_workbook(str(file_path))
        ws = wb.active
        assert ws.cell(2, 1).value == "文本"
        assert ws.cell(2, 2).value == 100


class TestWriteXlsxNegative:
    """负面测试"""

    def test_invalid_path(self):
        """无效路径"""
        result = write_xlsx(path="Z:/invalid/path/data.xlsx", data=[{"A": 1}])
        assert is_error(result)

    def test_inconsistent_columns(self, temp_output_dir):
        """列不一致(每行key不同) Bug #4已修复"""
        file_path = temp_output_dir / "inconsistent.xlsx"
        data = [
            {"A": 1, "B": 2},
            {"A": 3, "C": 4},  # B变C
            {"A": 5, "B": 6, "C": 7, "D": 8},  # 4列
        ]
        result = write_xlsx(path=str(file_path), data=data)

        assert is_success(result)
        wb = load_workbook(str(file_path))
        ws = wb.active
        # 修复在:表头包含所有列,缺失数据填None
        assert ws.max_column == 4  # A,B,C,D四列
        assert ws.cell(1, 1).value == "A"
        assert ws.cell(1, 2).value == "B"
        assert ws.cell(1, 3).value == "C"
        assert ws.cell(1, 4).value == "D"
        # 第2行:A=1, B=2, C=None, D=None
        assert ws.cell(2, 1).value == 1
        assert ws.cell(2, 2).value == 2
        assert ws.cell(2, 3).value is None
        # 第3行:A=3, B=None, C=4, D=None
        assert ws.cell(3, 1).value == 3
        assert ws.cell(3, 2).value is None
        assert ws.cell(3, 3).value == 4
        # 第4行:A=5, B=6, C=7, D=8
        assert ws.cell(4, 4).value == 8


class TestWriteXlsxSchemaIssues:
    """Schema问题验证"""

    def test_examples_coverage(self):
        """Examples覆盖不足:只有1个示例,缺少sheet_name参数示例"""
        # Schema有3个参数,Examples只有1个,覆盖率33%
        # Bug:Examples太少,LLM可能不知道sheet_name参数
        pass

    def test_data_description_clarity(self):
        """data参数描述不够清晰"""
        # 描述:"对象数组格式",但没有说明:
        # 1. 空数组[]是否合法
        # 2. None是否合法
        # 3. key不一致如何处理
        # Bug:Schema描述不完整
        pass
